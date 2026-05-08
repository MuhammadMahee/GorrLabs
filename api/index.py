from flask import Flask, render_template, request, jsonify, send_from_directory
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=BASE_DIR,
    static_folder=BASE_DIR,
    static_url_path="/static"
)

# ----------------------------
# FILE PATH (EXCEL)
# ----------------------------
EXCEL_FILE = os.path.join(BASE_DIR, "feedback.xlsx")


def init_excel():
    """Create Excel file if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Role", "Rating", "Message", "Time"])
        wb.save(EXCEL_FILE)


init_excel()

# ----------------------------
# SITEMAP
# ----------------------------
@app.route('/sitemap.xml')
def sitemap():
    return send_from_directory(BASE_DIR, 'sitemap.xml')


# ----------------------------
# EMAIL CONFIG
# ----------------------------
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = "gorrlabs@gmail.com"
SMTP_PASSWORD = "oyxt vvxv rmvw hsvg"
RECIPIENT = "gorrlabs@gmail.com"


def send_email(name, email, message):
    msg = MIMEMultipart("alternative")

    msg["Subject"] = "Gorr Labs — We received your message"
    msg["From"] = SMTP_USER
    msg["To"] = email
    msg["Cc"] = RECIPIENT

    recipients = [email, RECIPIENT]

    html = f"""
    <html>
    <body style="font-family:Arial,sans-serif;background:#0A0A0A;color:#fff;padding:30px;">
      <div style="max-width:600px;margin:auto;background:#111;border:1px solid #FF7A00;
                  border-radius:12px;padding:30px;text-align:center;">

        <img src="https://gorrlabs.vercel.app/static/logo.png"
             style="width:120px;margin-bottom:20px;">

        <h2 style="color:#FF7A00;">Gorr Labs</h2>

        <p style="text-align:left;">Hi {name},</p>

        <p style="text-align:left;">
          We’ve received your message and will get back to you within 24 hours.
        </p>

        <hr style="border-color:#333;margin:20px 0;">

        <p style="text-align:left;"><strong>Your Message:</strong></p>

        <p style="background:#1a1a1a;padding:16px;border-left:3px solid #FF2D2D;
                  border-radius:6px;text-align:left;">
          {message}
        </p>

        <hr style="border-color:#333;margin:20px 0;">

        <p style="font-size:12px;color:#777;text-align:left;">
          Submitted Email: {email}
        </p>

      </div>
    </body>
    </html>
    """

    msg.attach(MIMEText(html, "html"))

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASSWORD)
        s.sendmail(SMTP_USER, recipients, msg.as_string())


# ----------------------------
# ROUTES
# ----------------------------
@app.route("/")
def home():
    return render_template("index.html")


@app.route("/contact")
def contact():
    return render_template("contact.html")


@app.route("/feedback")
def feedback():
    return render_template("feedback.html")


@app.route("/view-feedback")
def view_feedback():
    return render_template("view-feedback.html")


# ----------------------------
# FEEDBACK SAVE (EXCEL)
# ----------------------------
@app.route('/submit-feedback', methods=['POST'])
def submit_feedback():

    data = request.get_json(silent=True) or {}

    row = [
        data.get("name", ""),
        data.get("role", "N/A"),
        data.get("rating", "5"),
        data.get("message", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ]

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        ws.append(row)
        wb.save(EXCEL_FILE)

        return jsonify({"success": True})

    except Exception as e:
        print("Excel Error:", e)
        return jsonify({"success": False, "error": str(e)}), 500


# ----------------------------
# FEEDBACK READ (EXCEL)
# ----------------------------
@app.route("/feedback-data")
def feedback_data():

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    data = []

    for r in rows[1:]:  # skip header
        data.append({
            "name": r[0],
            "role": r[1],
            "rating": r[2],
            "message": r[3],
            "time": r[4]
        })

    return jsonify({
        "success": True,
        "data": data[::-1]
    })


# ----------------------------
# CONTACT MESSAGE API
# ----------------------------
@app.route("/send-message", methods=["POST"])
def send_message():

    data = request.get_json(silent=True) or {}

    name = data.get("name", "").strip()
    email = data.get("email", "").strip()
    message = data.get("message", "").strip()

    if not name or not email or not message:
        return jsonify({
            "success": False,
            "error": "All fields are required."
        }), 400

    try:
        send_email(name, email, message)
        return jsonify({
            "success": True,
            "message": "Message sent successfully!"
        })

    except Exception as e:
        print("Email error:", e)
        return jsonify({
            "success": True,
            "message": "Message received (email fallback)."
        })


# ----------------------------
# RUN
# ----------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
